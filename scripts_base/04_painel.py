# ===============================================
# 🎧 SPIN Analyzer — Painel
# ✅ Saída: SOMENTE PLANILHAS
# ✅ Individual: exibe planilha na tela e permite download
# ✅ Gerencial: destaca o consolidado na tela e permite download do lote e individuais
# ✅ UX: visual premium, mensagens curtas, foco no cliente
# ✅ Progresso e tempo: atualizam durante o processamento
# ✅ Sessão leve: nunca guarda ZIP nem mapas grandes
# ✅ Limites no gerencial: Áudio até 5 arquivos e 10 minutos cada | Texto até 8 entradas
# ===============================================

import os
import re
import io
import time
import zipfile
import wave
import threading
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, List, Dict, Any

import streamlit as st
import requests


# ==============================
# ⚙️ set_page_config PRIMEIRO
# ==============================
st.set_page_config(
    page_title="SPIN Analyzer",
    page_icon="🎧",
    layout="wide",
)


# ==============================
# 🔐 Configurações
# ==============================
def _get_cfg(key: str, default: str = "") -> str:
    v = os.getenv(key)
    if v is not None:
        return str(v).strip()
    try:
        if key in st.secrets:
            return str(st.secrets[key]).strip()
    except Exception:
        pass
    return str(default).strip()


MODE = _get_cfg("MODE", "VPS").upper()
VPS_BASE_URL = _get_cfg("VPS_BASE_URL", "").rstrip("/")
VPS_API_KEY = _get_cfg("VPS_API_KEY", "")

CONNECT_TIMEOUT_S = int(_get_cfg("CONNECT_TIMEOUT_S", "10"))
READ_TIMEOUT_S = int(_get_cfg("API_TIMEOUT_S", "7200"))
REQ_TIMEOUT = (CONNECT_TIMEOUT_S, READ_TIMEOUT_S)

EXCEL_WRAP_TEXT = _get_cfg("EXCEL_WRAP_TEXT", "1").strip() not in ("0", "false", "False", "")
EXCEL_DEFAULT_COL_W = int(_get_cfg("EXCEL_DEFAULT_COL_W", "22"))
EXCEL_TEXT_COL_W = int(_get_cfg("EXCEL_TEXT_COL_W", "55"))
EXCEL_MAX_COL_W = int(_get_cfg("EXCEL_MAX_COL_W", "80"))

# Limites obrigatórios do gerencial
BATCH_WAV_MAX_FILES = 5
BATCH_WAV_MAX_SECONDS_EACH = 10 * 60
BATCH_TEXT_MAX_ENTRIES = 8

if MODE != "VPS":
    st.error("Este painel está configurado para funcionar apenas no modo online. Verifique as configurações do projeto.")
    st.stop()
if not VPS_BASE_URL:
    st.error("Configuração ausente. Endereço do serviço de avaliação.")
    st.stop()
if not VPS_API_KEY:
    st.error("Configuração ausente. Chave de acesso do serviço de avaliação.")
    st.stop()


# ==============================
# 🎨 Estilo premium
# ==============================
st.markdown(
    """
<style>
:root{
  --bg:#FFFFFF;
  --ink:#0B1220;
  --muted:#3A4A63;
  --line:rgba(199,214,245,0.95);
  --line2:rgba(199,214,245,0.70);
  --brand:#0B63F3;
  --brand2:#164DD6;
  --success:#29B37C;
  --soft:#F6F9FF;
}

body { background:var(--bg); color:var(--ink); font-family:Segoe UI, Arial, sans-serif; }
.block-container { padding-top: 1.1rem; padding-bottom: 2.0rem; max-width: 1500px; }
h1,h2,h3 { color:var(--brand); letter-spacing:0.2px; }

hr { border-color: var(--line2); }

.card{
  background:#FFFFFF !important;
  border:1px solid var(--line) !important;
  border-radius:18px;
  padding:18px;
  margin-bottom:14px;
  box-shadow:0 10px 28px rgba(11,18,32,0.08);
}

.card.tight{ padding:14px 16px; }

.smallmuted{ color:var(--muted); font-weight:650; }
.mini{ font-size:0.92rem; }

.hero{
  display:flex; align-items:flex-start; justify-content:space-between; gap:14px;
  padding:18px; border-radius:22px;
  border:1px solid rgba(199,214,245,0.85);
  background:linear-gradient(135deg, rgba(246,249,255,1) 0%, rgba(255,255,255,1) 55%, rgba(230,255,243,0.50) 100%);
  box-shadow:0 14px 34px rgba(11,18,32,0.10);
  margin-top:6px; margin-bottom:14px;
}
.hero .left{ display:flex; gap:12px; align-items:flex-start; }
.hero .icon{
  width:42px; height:42px; border-radius:14px;
  background:#EAF2FF; border:1px solid rgba(11,99,243,0.25);
  display:flex; align-items:center; justify-content:center; font-size:22px;
}
.hero .title{ margin:0; font-size:1.45rem; font-weight:900; color:var(--ink); }
.hero .subtitle{ margin:6px 0 0 0; color:var(--muted); font-weight:650; max-width: 840px; }

.pill-row{ display:flex; flex-wrap:wrap; gap:10px; align-items:center; justify-content:flex-end; }
.pill{
  display:inline-flex; align-items:center; gap:8px;
  padding:8px 12px; border-radius:999px;
  border:1px solid #AFC7F3; background:var(--soft);
  color:var(--brand); font-weight:800; font-size:0.92rem; line-height:1;
}
.pill .dot{ width:9px; height:9px; border-radius:999px; background:var(--brand); display:inline-block; }
.pill.ok{ background:#E6FFF3; border-color:var(--success); color:#0B6B4B; }
.pill.ok .dot{ background:var(--success); }

.status-card{
  background:linear-gradient(135deg, #FFFFFF 0%, var(--soft) 55%, #FFFFFF 100%) !important;
  border:1px solid var(--line) !important;
  border-radius:18px !important;
  padding:16px 18px !important;
  box-shadow:0 12px 30px rgba(11,18,32,0.08);
}
.status-title{ margin:0; font-size:1.10rem; font-weight:900; color:var(--ink); }
.status-sub{ margin:6px 0 0 0; color:var(--muted); font-weight:650; }

.section-title{
  margin: 0 0 6px 0;
  font-size: 1.05rem;
  font-weight: 900;
  color:var(--ink);
}

.kpi{
  display:flex; align-items:center; justify-content:space-between; gap:14px;
  padding:12px 14px; border-radius:16px;
  border:1px solid var(--line);
  background:#FFFFFF;
}
.kpi .k{ color:var(--muted); font-weight:750; }
.kpi .v{ color:var(--ink); font-weight:950; }

.table-shell{
  border:1px solid var(--line);
  background:#FFFFFF;
  border-radius:18px;
  padding:12px;
  box-shadow:0 10px 24px rgba(11,18,32,0.06);
}

.hint{
  border-left:4px solid rgba(11,99,243,0.35);
  background:rgba(246,249,255,0.65);
  padding:10px 12px;
  border-radius:12px;
  color:var(--muted);
  font-weight:650;
}

button[kind="primary"]{
  border-radius:14px !important;
}

[data-testid="stDownloadButton"] button{
  border-radius:14px !important;
  padding:0.72rem 0.95rem !important;
  font-weight:850 !important;
}

[data-testid="stButton"] button{
  border-radius:14px !important;
  padding:0.72rem 0.95rem !important;
  font-weight:850 !important;
}

[data-testid="stDataFrame"]{
  border-radius:14px;
}

</style>
""",
    unsafe_allow_html=True,
)


# ==============================
# 🧠 Estado leve
# ==============================
def _ensure_state():
    st.session_state.setdefault("view", "single")
    st.session_state.setdefault("single_mode", "Texto")
    st.session_state.setdefault("batch_mode", "Texto")

    st.session_state.setdefault("last_result", None)
    st.session_state.setdefault("batch_results", None)
    st.session_state.setdefault("batch_lote", None)

    st.session_state.setdefault("_prev_view", st.session_state["view"])
    st.session_state.setdefault("_prev_single_mode", st.session_state["single_mode"])
    st.session_state.setdefault("_prev_batch_mode", st.session_state["batch_mode"])


_ensure_state()


def clear_all_results():
    st.session_state["last_result"] = None
    st.session_state["batch_results"] = None
    st.session_state["batch_lote"] = None


def clear_single():
    st.session_state["last_result"] = None


def clear_batch():
    st.session_state["batch_results"] = None
    st.session_state["batch_lote"] = None


# ==============================
# ✅ Validação do texto
# ==============================
def validar_transcricao(txt: str) -> Tuple[bool, str]:
    linhas = [l.strip() for l in (txt or "").splitlines() if l.strip()]
    if len(linhas) < 4:
        return False, "O conteúdo está muito curto para avaliação."
    if not any(re.match(r"^\[(VENDEDOR|CLIENTE)\]", l, re.I) for l in linhas):
        return False, "Formato inválido. Comece as falas com [VENDEDOR] e [CLIENTE]."
    return True, "ok"


# ==============================
# ⏱️ Utilidades
# ==============================
def duracao_wav_seg_bytes(wav_bytes: bytes) -> float:
    try:
        bio = io.BytesIO(wav_bytes)
        with wave.open(bio, "rb") as wf:
            return wf.getnframes() / float(wf.getframerate())
    except Exception:
        return 0.0


def human_time(sec: float) -> str:
    try:
        sec = float(sec)
    except Exception:
        sec = 0.0
    if sec < 60:
        return f"{int(sec)}s"
    return f"{int(sec // 60)}m {int(sec % 60)}s"


# ==============================
# 📏 Excel: formatar largura e wrap
# ==============================
def format_excel_bytes(excel_bytes: bytes) -> bytes:
    if not excel_bytes:
        return excel_bytes
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
    except Exception:
        return excel_bytes

    bio = io.BytesIO(excel_bytes)
    wb = load_workbook(bio)
    ws = wb.active
    ws.freeze_panes = "A2"

    long_text_markers = ("_texto", "_feedback", "justific", "trecho", "observ", "coment", "resumo", "explic")

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col_idx).value
        if v is not None:
            headers[str(v)] = col_idx

    wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    normal_align = Alignment(wrap_text=False, vertical="top", horizontal="left")
    max_rows = min(ws.max_row, 5000)

    for header, col_idx in headers.items():
        h = str(header).lower()
        is_long = any(m in h for m in long_text_markers)

        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = min(
            EXCEL_TEXT_COL_W if is_long else EXCEL_DEFAULT_COL_W,
            EXCEL_MAX_COL_W,
        )

        if EXCEL_WRAP_TEXT:
            for r in range(1, max_rows + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.alignment = wrap_align if is_long else normal_align

    ws.row_dimensions[1].height = 22
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ==============================
# 📊 Excel: leitura para visualização
# ==============================
def _excel_to_dataframe(excel_bytes: bytes):
    if not excel_bytes:
        return None

    bio = io.BytesIO(excel_bytes)

    # Preferência: pandas, por ser mais estável para leitura rápida
    try:
        import pandas as pd
        df = pd.read_excel(bio, sheet_name=0, engine="openpyxl")
        if df is None:
            return None
        # Garante colunas como texto para melhor leitura na tabela
        df.columns = [str(c) for c in df.columns]
        return df
    except Exception:
        pass

    # Fallback: openpyxl sem pandas
    try:
        from openpyxl import load_workbook
        wb = load_workbook(bio, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None
        headers = [str(h) if h is not None else "" for h in rows[0]]
        data = rows[1:]
        # Evita depender de pandas, devolve estrutura simples
        return {"headers": headers, "data": data}
    except Exception:
        return None


def _infer_column_config_from_df(df, sample_rows: int = 180):
    """
    Heurísticas simples para dar prioridade de largura a colunas de texto.
    Mantém consistência sem dependências externas.
    """
    col_cfg = {}
    if df is None or not hasattr(df, "columns"):
        return col_cfg

    try:
        sample = df.head(sample_rows)
    except Exception:
        sample = df

    long_markers = ("texto", "feedback", "justific", "trecho", "observ", "coment", "resumo", "explic", "descricao", "mensagem")
    id_markers = ("id", "codigo", "protocolo", "arquivo", "nome")

    for c in list(df.columns):
        c_str = str(c)
        c_low = c_str.lower()

        # Medida simples de tamanho com base no cabeçalho e amostra
        try:
            vals = sample[c].astype(str).fillna("").tolist()
            max_len = max([len(c_str)] + [len(v) for v in vals[:sample_rows]])
        except Exception:
            max_len = len(c_str)

        is_long = any(m in c_low for m in long_markers) or max_len >= 45
        is_id = any(m in c_low for m in id_markers) and max_len <= 18
        is_short = max_len <= 10

        if is_long:
            col_cfg[c_str] = st.column_config.TextColumn(label=c_str, width="large")
        elif is_id:
            col_cfg[c_str] = st.column_config.TextColumn(label=c_str, width="small")
        elif is_short:
            col_cfg[c_str] = st.column_config.TextColumn(label=c_str, width="small")
        else:
            col_cfg[c_str] = st.column_config.TextColumn(label=c_str, width="medium")

    return col_cfg


def render_excel_open(excel_bytes: bytes, title: str, subtitle: str, height: int = 620):
    st.markdown(
        f"""
<div class="card">
  <div class="section-title">{title}</div>
  <div class="smallmuted">{subtitle}</div>
</div>
""",
        unsafe_allow_html=True,
    )

    df_or = _excel_to_dataframe(excel_bytes)

    # Fallback sem pandas
    if isinstance(df_or, dict) and "headers" in df_or:
        headers = df_or["headers"]
        data = df_or["data"]
        st.markdown("<div class='table-shell'>", unsafe_allow_html=True)
        st.dataframe(
            data,
            use_container_width=True,
            height=height,
        )
        st.markdown("</div>", unsafe_allow_html=True)
        return

    df = df_or
    if df is None:
        st.info("Não foi possível abrir a planilha para visualização. O download continua disponível.")
        return

    # Ajustes de leitura
    try:
        df = df.copy()
        df.columns = [str(c) for c in df.columns]
        # Evita colunas totalmente vazias no final, quando existir
        empty_cols = [c for c in df.columns if df[c].isna().all()]
        if empty_cols and len(empty_cols) < len(df.columns):
            df = df.drop(columns=empty_cols)
    except Exception:
        pass

    col_cfg = _infer_column_config_from_df(df)

    st.markdown("<div class='table-shell'>", unsafe_allow_html=True)
    st.dataframe(
        df,
        use_container_width=True,
        height=height,
        column_config=col_cfg if col_cfg else None,
        hide_index=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)


# ==============================
# 🌐 Saúde do serviço
# ==============================
def vps_health() -> bool:
    try:
        r = requests.get(f"{VPS_BASE_URL}/health", timeout=(3, 6))
        return r.status_code == 200
    except Exception:
        return False


# ==============================
# 📦 ZIP helpers
# ==============================
def zip_extract_all(zip_bytes: bytes) -> Dict[str, bytes]:
    out: Dict[str, bytes] = {}
    bio = io.BytesIO(zip_bytes)
    with zipfile.ZipFile(bio, "r") as z:
        for name in z.namelist():
            try:
                out[name] = z.read(name)
            except Exception:
                pass
    return out


def pick_excels(files_map: Dict[str, bytes]) -> List[Tuple[str, bytes]]:
    excels = [(k, v) for k, v in files_map.items() if k.lower().endswith(".xlsx")]
    if not excels:
        return []

    def _score(name: str) -> int:
        n = name.lower()
        if "spin_resultados_lote" in n:
            return 0
        if n.endswith("_spin.xlsx") or "_spin" in n:
            return 1
        return 2

    excels.sort(key=lambda kv: (_score(kv[0]), kv[0]))
    return excels


# ==============================
# ✅ UI helpers
# ==============================
def render_hero(title: str, subtitle: str, pills: List[str], icon: str):
    st.markdown(
        f"""
<div class="hero">
  <div class="left">
    <div class="icon">{icon}</div>
    <div>
      <p class="title">{title}</p>
      <p class="subtitle">{subtitle}</p>
    </div>
  </div>
  <div class="pill-row">
    {''.join(pills)}
  </div>
</div>
""",
        unsafe_allow_html=True,
    )


def show_friendly_error(title: str, err: Exception):
    st.error(title)
    with st.expander("Detalhes do diagnóstico", expanded=False):
        st.code(str(err))


# ==============================
# ⏳ Progresso e tempo que atualizam
# ==============================
def run_with_live_progress(task_fn, phase_labels: List[str]):
    """
    task_fn: função sem streamlit que executa e retorna resultado
    phase_labels: mensagens curtas para o cliente
    """
    status = st.empty()
    timer = st.empty()
    pbar = st.progress(0)

    start = time.time()
    done_flag = {"ok": False}
    result_box = {"value": None}
    error_box = {"error": None}

    def _worker():
        try:
            result_box["value"] = task_fn()
            done_flag["ok"] = True
        except Exception as e:
            error_box["error"] = e
            done_flag["ok"] = True

    th = threading.Thread(target=_worker, daemon=True)
    th.start()

    # Progresso indeterminado mais estável e um pouco mais lento
    # Sobe até 30, depois avança gradualmente até 90 enquanto aguarda
    phase = 0
    last_phase_change = 0.0

    while not done_flag["ok"]:
        elapsed = time.time() - start

        timer.markdown(
            f"<div class='smallmuted'>Tempo decorrido <b>{human_time(elapsed)}</b></div>",
            unsafe_allow_html=True,
        )

        if elapsed - last_phase_change > 5.2 and phase < len(phase_labels) - 1:
            phase += 1
            last_phase_change = elapsed

        msg = phase_labels[min(phase, len(phase_labels) - 1)]
        status.markdown(
            f"""
<div class="status-card">
  <p class="status-title">{msg}</p>
  <p class="status-sub">Acompanhe o andamento em tempo real</p>
</div>
""",
            unsafe_allow_html=True,
        )

        if elapsed < 10:
            prog = int(4 + (elapsed / 10) * 26)  # 4..30
        else:
            # aproxima de 90 sem travar e com avanço mais suave
            prog = int(30 + (1 - (1 / (1 + (elapsed - 10) / 11))) * 60)  # 30..90

        prog = max(0, min(90, prog))
        pbar.progress(prog)

        time.sleep(0.22)

    elapsed = time.time() - start
    timer.markdown(
        f"<div class='smallmuted'>Tempo decorrido <b>{human_time(elapsed)}</b></div>",
        unsafe_allow_html=True,
    )
    status.markdown(
        """
<div class="status-card">
  <p class="status-title">Finalização e conferência</p>
  <p class="status-sub">A planilha será exibida em seguida</p>
</div>
""",
        unsafe_allow_html=True,
    )
    pbar.progress(100)
    time.sleep(0.35)

    status.empty()
    timer.empty()
    pbar.empty()

    if error_box["error"] is not None:
        raise error_box["error"]

    return result_box["value"], float(elapsed)


# ==============================
# 🌐 Chamada ao serviço remoto
# ==============================
def vps_run_file_blocking(file_bytes: bytes, filename: str, mime: str) -> Tuple[bytes, Dict[str, str]]:
    files = {"file": (filename, file_bytes, mime)}
    headers = {"X-API-KEY": VPS_API_KEY}

    r = requests.post(
        f"{VPS_BASE_URL}/run",
        files=files,
        headers=headers,
        timeout=REQ_TIMEOUT,
    )
    r.raise_for_status()
    zip_bytes = r.content
    useful = {"X-Run-Id": r.headers.get("X-Run-Id", "")}
    return zip_bytes, useful


# ==============================
# ✅ Processamento Individual
# ==============================
def run_single_text(content: str):
    ok, msg = validar_transcricao(content)
    if not ok:
        st.error(msg)
        return

    fname = f"avaliacao_texto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    def task():
        return vps_run_file_blocking(content.encode("utf-8", errors="ignore"), fname, "text/plain")

    try:
        (zip_bytes, hdr), elapsed = run_with_live_progress(
            task_fn=task,
            phase_labels=["Preparação do material", "Avaliação em andamento", "Geração da planilha", "Organização dos resultados"],
        )
    except Exception as e:
        show_friendly_error("Não foi possível concluir a avaliação.", e)
        return

    files_map = zip_extract_all(zip_bytes)
    excels = pick_excels(files_map)
    if not excels:
        st.error("Recebi um retorno, mas não encontrei nenhuma planilha.")
        with st.expander("Arquivos retornados", expanded=False):
            st.write(list(files_map.keys())[:200])
        return

    main_name, main_xlsx = excels[0]
    main_xlsx_fmt = format_excel_bytes(main_xlsx)

    st.session_state["last_result"] = {
        "kind": "texto",
        "run_id": hdr.get("X-Run-Id", ""),
        "filename": fname,
        "excel_name": main_name,
        "excel_bytes": main_xlsx_fmt,
        "timings": {"audio_sec": 0.0, "total_sec": float(elapsed)},
    }


def run_single_audio(wav_file):
    wav_bytes = wav_file.getbuffer().tobytes()
    audio_sec = duracao_wav_seg_bytes(wav_bytes)
    if audio_sec and audio_sec > BATCH_WAV_MAX_SECONDS_EACH:
        st.error(f"Este áudio tem cerca de {audio_sec/60:.1f} minutos. Para melhor experiência, use até 10 minutos.")
        return

    def task():
        return vps_run_file_blocking(wav_bytes, wav_file.name, "audio/wav")

    try:
        (zip_bytes, hdr), elapsed = run_with_live_progress(
            task_fn=task,
            phase_labels=["Preparação do material", "Transcrição e avaliação", "Geração da planilha", "Organização dos resultados"],
        )
    except Exception as e:
        show_friendly_error("Não foi possível concluir a avaliação.", e)
        return

    files_map = zip_extract_all(zip_bytes)
    excels = pick_excels(files_map)
    if not excels:
        st.error("Recebi um retorno, mas não encontrei nenhuma planilha.")
        with st.expander("Arquivos retornados", expanded=False):
            st.write(list(files_map.keys())[:200])
        return

    main_name, main_xlsx = excels[0]
    main_xlsx_fmt = format_excel_bytes(main_xlsx)

    st.session_state["last_result"] = {
        "kind": "áudio",
        "run_id": hdr.get("X-Run-Id", ""),
        "filename": wav_file.name,
        "excel_name": main_name,
        "excel_bytes": main_xlsx_fmt,
        "timings": {"audio_sec": float(audio_sec or 0.0), "total_sec": float(elapsed)},
    }


# ==============================
# ✅ Processamento Gerencial
# ==============================
def run_batch_text(files: List[Any], pasted_blocks: List[str]):
    entradas: List[Tuple[str, str]] = []

    if files:
        for f in files:
            try:
                content = f.getvalue().decode("utf-8", errors="ignore")
            except Exception:
                content = ""
            entradas.append((f.name, content))

    if pasted_blocks:
        for i, b in enumerate(pasted_blocks, start=1):
            entradas.append((f"colado_{i}.txt", b))

    entradas = [(n, t) for (n, t) in entradas if (t or "").strip()]

    if not entradas:
        st.error("Envie arquivos ou cole pelo menos um conteúdo.")
        return

    if len(entradas) > BATCH_TEXT_MAX_ENTRIES:
        st.error(f"No gerencial, use até {BATCH_TEXT_MAX_ENTRIES} entradas por vez.")
        return

    for name, txt in entradas:
        ok, msg = validar_transcricao(txt)
        if not ok:
            st.error(f"{name}: {msg}")
            return

    itens: List[dict] = []
    lote_excel_bytes = b""
    lote_excel_name = ""

    total = len(entradas)

    for idx, (name, txt) in enumerate(entradas, start=1):
        def task():
            return vps_run_file_blocking(txt.encode("utf-8", errors="ignore"), name, "text/plain")

        try:
            (zip_bytes, hdr), _elapsed = run_with_live_progress(
                task_fn=task,
                phase_labels=[
                    f"Preparação do item {idx} de {total}",
                    f"Avaliação do item {idx} de {total}",
                    "Geração da planilha",
                    "Organização dos resultados",
                ],
            )
        except Exception as e:
            show_friendly_error(f"Não foi possível concluir o item {idx}.", e)
            return

        files_map = zip_extract_all(zip_bytes)
        excels = pick_excels(files_map)

        indiv_name = ""
        indiv_xlsx_fmt = b""
        if excels:
            chosen = None
            for nm, xb in excels:
                if nm.lower().endswith("_spin.xlsx") or "_spin" in nm.lower():
                    chosen = (nm, xb)
                    break
            if not chosen:
                chosen = excels[0]
            indiv_name, indiv_xlsx = chosen
            indiv_xlsx_fmt = format_excel_bytes(indiv_xlsx)

        itens.append(
            {
                "idx": idx,
                "filename": name,
                "run_id": hdr.get("X-Run-Id", ""),
                "excel_individual_name": indiv_name,
                "excel_individual_bytes": indiv_xlsx_fmt,
            }
        )

        for nm, xb in excels:
            if "spin_resultados_lote" in nm.lower():
                lote_excel_name = nm
                lote_excel_bytes = format_excel_bytes(xb)
                break

    st.session_state["batch_results"] = itens
    st.session_state["batch_lote"] = {
        "excel_name": lote_excel_name,
        "excel_bytes": lote_excel_bytes,
    } if lote_excel_bytes else None


def run_batch_audio(wavs: List[Any]):
    if not wavs:
        st.error("Envie pelo menos 1 áudio.")
        return

    if len(wavs) > BATCH_WAV_MAX_FILES:
        st.error(f"No gerencial, envie até {BATCH_WAV_MAX_FILES} áudios por vez.")
        return

    for wf in wavs:
        b = wf.getbuffer().tobytes()
        d = duracao_wav_seg_bytes(b)
        if d and d > BATCH_WAV_MAX_SECONDS_EACH:
            st.error(f"{wf.name} tem cerca de {d/60:.1f} minutos. No gerencial, use até 10 minutos por áudio.")
            return

    itens: List[dict] = []
    lote_excel_bytes = b""
    lote_excel_name = ""

    total = len(wavs)

    for idx, wavf in enumerate(wavs, start=1):
        wav_bytes = wavf.getbuffer().tobytes()

        def task():
            return vps_run_file_blocking(wav_bytes, wavf.name, "audio/wav")

        try:
            (zip_bytes, hdr), _elapsed = run_with_live_progress(
                task_fn=task,
                phase_labels=[
                    f"Preparação do item {idx} de {total}",
                    f"Transcrição e avaliação do item {idx} de {total}",
                    "Geração da planilha",
                    "Organização dos resultados",
                ],
            )
        except Exception as e:
            show_friendly_error(f"Não foi possível concluir o item {idx}.", e)
            return

        files_map = zip_extract_all(zip_bytes)
        excels = pick_excels(files_map)

        indiv_name = ""
        indiv_xlsx_fmt = b""
        if excels:
            chosen = None
            for nm, xb in excels:
                if nm.lower().endswith("_spin.xlsx") or "_spin" in nm.lower():
                    chosen = (nm, xb)
                    break
            if not chosen:
                chosen = excels[0]
            indiv_name, indiv_xlsx = chosen
            indiv_xlsx_fmt = format_excel_bytes(indiv_xlsx)

        itens.append(
            {
                "idx": idx,
                "filename": wavf.name,
                "run_id": hdr.get("X-Run-Id", ""),
                "excel_individual_name": indiv_name,
                "excel_individual_bytes": indiv_xlsx_fmt,
            }
        )

        for nm, xb in excels:
            if "spin_resultados_lote" in nm.lower():
                lote_excel_name = nm
                lote_excel_bytes = format_excel_bytes(xb)
                break

    st.session_state["batch_results"] = itens
    st.session_state["batch_lote"] = {
        "excel_name": lote_excel_name,
        "excel_bytes": lote_excel_bytes,
    } if lote_excel_bytes else None


# ==============================
# Cabeçalho
# ==============================
render_hero(
    title="SPIN Analyzer",
    subtitle="Avaliação em planilha pronta para acompanhamento executivo e auditoria.",
    pills=[
        "<span class='pill ok'><span class='dot'></span>Planilha em tela</span>",
        "<span class='pill'><span class='dot'></span>Individual e Gerencial</span>",
    ],
    icon="🎧",
)
st.markdown("---")


# ==============================
# Sidebar
# ==============================
with st.sidebar:
    st.markdown("### Navegação")

    st.session_state["view"] = st.radio(
        "Área",
        options=["single", "batch"],
        format_func=lambda v: "Avaliação individual" if v == "single" else "Visão gerencial",
        key="view_radio",
    )

    if st.session_state["view"] != st.session_state["_prev_view"]:
        clear_all_results()
        st.session_state["_prev_view"] = st.session_state["view"]

    st.markdown("---")

    if vps_health():
        st.success("Serviço disponível")
    else:
        st.warning("Serviço indisponível")

    st.markdown("---")
    if st.button("Limpar resultados", use_container_width=True):
        clear_all_results()


# ==============================
# Telas
# ==============================
if st.session_state["view"] == "single":
    st.markdown("### Avaliação individual")

    st.session_state["single_mode"] = st.radio(
        "Formato",
        options=["Texto", "Áudio"],
        horizontal=True,
        key="single_mode_radio",
    )

    if st.session_state["single_mode"] != st.session_state["_prev_single_mode"]:
        clear_single()
        st.session_state["_prev_single_mode"] = st.session_state["single_mode"]

    if st.session_state["single_mode"] == "Texto":
        st.markdown(
            "<div class='hint'>Marque as falas com <b>[VENDEDOR]</b> e <b>[CLIENTE]</b> para melhor precisão.</div>",
            unsafe_allow_html=True,
        )

        content = st.text_area(
            "Conteúdo para avaliação",
            height=260,
            value="",
            key="single_text",
            placeholder="[VENDEDOR] ...\n[CLIENTE] ...\n[VENDEDOR] ...",
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Iniciar avaliação", use_container_width=True, type="primary"):
                clear_single()
                run_single_text(content)

        with col2:
            if st.button("Limpar", use_container_width=True):
                clear_single()

    else:
        up = st.file_uploader(
            "Áudio para avaliação",
            type=["wav"],
            accept_multiple_files=False,
            key="single_audio",
        )

        st.markdown(
            "<div class='hint'>Para melhor experiência, priorize gravações com boa qualidade de áudio.</div>",
            unsafe_allow_html=True,
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Iniciar avaliação", use_container_width=True, type="primary"):
                if up is None:
                    st.error("Envie um áudio para continuar.")
                else:
                    clear_single()
                    run_single_audio(up)

        with col2:
            if st.button("Limpar", use_container_width=True):
                clear_single()

else:
    st.markdown("### Visão gerencial")

    st.session_state["batch_mode"] = st.radio(
        "Formato",
        options=["Texto", "Áudio"],
        horizontal=True,
        key="batch_mode_radio",
    )

    if st.session_state["batch_mode"] != st.session_state["_prev_batch_mode"]:
        clear_batch()
        st.session_state["_prev_batch_mode"] = st.session_state["batch_mode"]

    if st.session_state["batch_mode"] == "Texto":
        st.markdown(
            f"""
<div class="card tight">
  <div class="section-title">Limites</div>
  <div class="smallmuted">Até {BATCH_TEXT_MAX_ENTRIES} entradas por vez</div>
</div>
""",
            unsafe_allow_html=True,
        )

        st.markdown(
            "<div class='hint'>Marque as falas com <b>[VENDEDOR]</b> e <b>[CLIENTE]</b>.</div>",
            unsafe_allow_html=True,
        )

        up_files = st.file_uploader(
            "Arquivos de texto",
            type=["txt"],
            accept_multiple_files=True,
            key="batch_text_files",
        )

        st.markdown(
            "<div class='smallmuted mini'>Você pode colar vários blocos separados por uma linha com <b>---</b></div>",
            unsafe_allow_html=True,
        )
        multi = st.text_area(
            "Conteúdos colados",
            height=220,
            value="",
            key="batch_text_area",
            placeholder="[VENDEDOR] ...\n[CLIENTE] ...\n---\n[VENDEDOR] ...\n[CLIENTE] ...",
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Iniciar avaliação", use_container_width=True, type="primary"):
                blocks = []
                if multi.strip():
                    blocks = [b.strip() for b in multi.split("\n---\n") if b.strip()]
                clear_batch()
                run_batch_text(up_files or [], blocks)

        with col2:
            if st.button("Limpar", use_container_width=True):
                clear_batch()

    else:
        st.markdown(
            f"""
<div class="card tight">
  <div class="section-title">Limites</div>
  <div class="smallmuted">Até {BATCH_WAV_MAX_FILES} áudios por vez</div>
  <div class="smallmuted">Até 10 minutos por áudio</div>
</div>
""",
            unsafe_allow_html=True,
        )

        up_wavs = st.file_uploader(
            "Áudios para avaliação",
            type=["wav"],
            accept_multiple_files=True,
            key="batch_audio_files",
        )

        st.markdown(
            "<div class='hint'>O consolidado será exibido como destaque ao final do processamento.</div>",
            unsafe_allow_html=True,
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Iniciar avaliação", use_container_width=True, type="primary"):
                clear_batch()
                run_batch_audio(up_wavs or [])

        with col2:
            if st.button("Limpar", use_container_width=True):
                clear_batch()


# ==============================
# Resultado individual
# ==============================
lr = st.session_state.get("last_result")
if lr and lr.get("excel_bytes"):
    st.markdown("---")

    pills = []
    kind = lr.get("kind", "")
    if kind:
        pills.append(f"<span class='pill ok'><span class='dot'></span>{kind.upper()}</span>")
    if lr.get("run_id"):
        pills.append(f"<span class='pill'><span class='dot'></span>Protocolo {lr.get('run_id')}</span>")

    render_hero(
        title="Resultado",
        subtitle="A planilha está pronta e já está aberta abaixo para conferência.",
        pills=pills,
        icon="✅",
    )

    # KPIs discretos
    t = (lr.get("timings") or {})
    total_sec = float(t.get("total_sec") or 0.0)
    audio_sec = float(t.get("audio_sec") or 0.0)
    k1, k2, k3 = st.columns(3)
    with k1:
        st.markdown(
            f"<div class='kpi'><div class='k'>Tempo total</div><div class='v'>{human_time(total_sec)}</div></div>",
            unsafe_allow_html=True,
        )
    with k2:
        st.markdown(
            f"<div class='kpi'><div class='k'>Duração do áudio</div><div class='v'>{human_time(audio_sec) if audio_sec else '—'}</div></div>",
            unsafe_allow_html=True,
        )
    with k3:
        st.markdown(
            f"<div class='kpi'><div class='k'>Arquivo</div><div class='v'>{Path(lr.get('filename') or '').name or '—'}</div></div>",
            unsafe_allow_html=True,
        )

    # Planilha aberta na tela
    render_excel_open(
        excel_bytes=lr.get("excel_bytes", b""),
        title="Planilha aberta",
        subtitle="Visualização executiva com rolagem horizontal e vertical e prioridade para colunas de texto.",
        height=650,
    )

    st.markdown(
        """
<div class="card">
  <div class="section-title">Download</div>
  <div class="smallmuted">Baixe a planilha para arquivamento, auditoria e compartilhamento interno.</div>
</div>
""",
        unsafe_allow_html=True,
    )

    filename = lr.get("filename") or f"avaliacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    base = Path(filename).stem

    st.download_button(
        "Baixar planilha",
        data=lr.get("excel_bytes", b""),
        file_name=f"{base}_avaliacao.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"dl_single_{base}",
    )


# ==============================
# Resultado gerencial
# ==============================
br = st.session_state.get("batch_results")
batch_lote = st.session_state.get("batch_lote")

if br:
    st.markdown("---")

    render_hero(
        title="Resultados do lote",
        subtitle="O consolidado é o destaque e fica aberto para leitura imediata. As planilhas individuais permanecem disponíveis para download.",
        pills=[
            "<span class='pill ok'><span class='dot'></span>Consolidado em destaque</span>",
            "<span class='pill'><span class='dot'></span>Individuais para download</span>",
        ],
        icon="📊",
    )

    if batch_lote and batch_lote.get("excel_bytes"):
        # Planilha do lote aberta primeiro
        render_excel_open(
            excel_bytes=batch_lote["excel_bytes"],
            title="Planilha consolidada",
            subtitle="Visualização ampla e legível para análise do lote.",
            height=680,
        )

        st.markdown(
            """
<div class="card">
  <div class="section-title">Download do consolidado</div>
  <div class="smallmuted">Baixe a planilha do lote para registro e acompanhamento.</div>
</div>
""",
            unsafe_allow_html=True,
        )

        st.download_button(
            "Baixar planilha do lote",
            data=batch_lote["excel_bytes"],
            file_name=f"lote_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl_lote",
        )
    else:
        st.warning("O consolidado não ficou disponível neste retorno. As planilhas individuais continuam acessíveis.")

    st.markdown(
        """
<div class="card">
  <div class="section-title">Planilhas individuais</div>
  <div class="smallmuted">Use quando precisar auditar um item específico do lote.</div>
</div>
""",
        unsafe_allow_html=True,
    )

    for item in br:
        idx = item.get("idx", 0)
        filename = str(item.get("filename") or f"item_{idx}")
        base = Path(filename).stem

        with st.expander(f"{idx}. {filename}", expanded=False):
            xb = item.get("excel_individual_bytes", b"")
            if xb:
                st.download_button(
                    "Baixar planilha individual",
                    data=xb,
                    file_name=f"{base}_avaliacao.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"dl_item_{idx}_{base}",
                )
            else:
                st.warning("Planilha individual não disponível para este item.")


# ==============================
# Rodapé
# ==============================
st.markdown("---")
st.markdown(
    "<div style='text-align:center;color:#3A4A63;font-weight:650;'>SPIN Analyzer — Projeto Tele IA 2026 - Desenvolvido Por Pualo Coutinho</div>",
    unsafe_allow_html=True,
)
