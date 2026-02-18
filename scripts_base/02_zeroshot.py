#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
02_zeroshot.py — SPIN Analyzer (LOCAL) — v8.1.1 (Windows-friendly)

TXT -> Ollama -> TSV (tolerante) -> Excel individual.

Arquivos:
- Entrada (default): arquivos_transcritos/txt (recursivo por padrão)
- Saída (default): saida_excel
- Prompts (sempre lidos do arquivo):
  - assets/Command_Core_D_Check_V2_6.txt
  - assets/Command_Core_D_Check_V2_6_FALLBACK.txt
- Cache SQLite: cache_spin02/cache.db
- Arquivamento de TXT processado: cachebd/ (move após processar)

CLI:
--in_dir, --out_dir, --pattern, --recursive, --workers, --force, --quiet

Exit code:
0 se tudo OK
1 se houve falha em algum arquivo
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import re
import shutil
import sqlite3
import sys
import time
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib import request
from urllib.error import HTTPError, URLError

from openpyxl import Workbook


# ============================================================
# Constantes de Fases / Formato
# ============================================================

PHASES: List[str] = [
    "P0_abertura",
    "P1_situation",
    "P2_problem",
    "P3_implication",
    "P4_need_payoff",
]

# Header canônico (3 colunas). O parser agora aceita variantes e reconstrói este formato.
TSV_HEADER = "SPIN SELLING\tCHECK_01\tCHECK_02"

TAG_RE = re.compile(
    r"^\s*(\[(VENDEDOR|CLIENTE)\]|(VENDEDOR|CLIENTE|AGENTE|ATENDENTE)\s*:)\s*",
    re.IGNORECASE,
)


# ============================================================
# Paths (projeto)
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
ROOT_DIR = BASE_DIR.parent

DEFAULT_IN_DIR = ROOT_DIR / "arquivos_transcritos" / "txt"
DEFAULT_OUT_DIR = ROOT_DIR / "saida_excel"

ASSETS_DIR = ROOT_DIR / "assets"
PROMPT_MAIN_PATH = ASSETS_DIR / "Command_Core_D_Check_V2_6.txt"
PROMPT_ALT_PATH = ASSETS_DIR / "Command_Core_D_Check_V2_6_FALLBACK.txt"

LOG_DIR = ROOT_DIR / "logs"
LOG_FILE = LOG_DIR / "spin02.log"

CACHE_DIR_DEFAULT = ROOT_DIR / "cache_spin02"
CACHE_DB_PATH = CACHE_DIR_DEFAULT / "cache.db"

ARCHIVE_DIR = ROOT_DIR / "cachebd"  # pasta pedida para mover os TXT já processados


# ============================================================
# Config (env)
# ============================================================

def _env_str(name: str, default: str) -> str:
    v = os.getenv(name, "").strip()
    return v if v else default

def _env_int(name: str, default: int) -> int:
    v = os.getenv(name, "").strip()
    if not v:
        return default
    try:
        return int(v)
    except Exception:
        return default

def _env_float(name: str, default: float) -> float:
    v = os.getenv(name, "").strip()
    if not v:
        return default
    try:
        return float(v)
    except Exception:
        return default


OLLAMA_URL = _env_str("OLLAMA_URL", "http://127.0.0.1:11434/api/generate")
OLLAMA_MODEL = _env_str("OLLAMA_MODEL", "qwen2.5:14b-instruct-q4_K_M")
OLLAMA_TIMEOUT_S = _env_int("OLLAMA_TIMEOUT_S", 900)
OLLAMA_TIMEOUT_RETRIES = _env_int("OLLAMA_TIMEOUT_RETRIES", 1)
OLLAMA_KEEP_ALIVE = _env_str("OLLAMA_KEEP_ALIVE", "30m")

OLLAMA_TEMPERATURE = _env_float("OLLAMA_TEMPERATURE", 0.0)
OLLAMA_NUM_CTX = _env_int("OLLAMA_NUM_CTX", 2048)
OLLAMA_NUM_PREDICT = _env_int("OLLAMA_NUM_PREDICT", 220)
OLLAMA_TOP_P = _env_float("OLLAMA_TOP_P", 0.9)
OLLAMA_REPEAT_PENALTY = _env_float("OLLAMA_REPEAT_PENALTY", 1.06)

SPIN_VENDOR_ONLY = (_env_str("SPIN_VENDOR_ONLY", "1") != "0")
SPIN_MAX_LINES_TOTAL = _env_int("SPIN_MAX_LINES_TOTAL", 260)
SPIN_MAX_CHARS_TOTAL = _env_int("SPIN_MAX_CHARS_TOTAL", 12000)

# Heartbeat enquanto o Ollama roda (para mostrar que não travou)
HEARTBEAT_EVERY_S = _env_int("SPIN_HEARTBEAT_EVERY_S", 25)


# ============================================================
# Logging
# ============================================================

def setup_logging(quiet: bool) -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger("spin02")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    fmt_file = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    fh = logging.FileHandler(str(LOG_FILE), encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(fmt_file)
    logger.addHandler(fh)

    ch = logging.StreamHandler(stream=sys.stdout)
    ch.setLevel(logging.ERROR if quiet else logging.INFO)
    ch.setFormatter(logging.Formatter(fmt="%(message)s"))
    logger.addHandler(ch)

    return logger


# ============================================================
# Util
# ============================================================

def sha256_text(s: str) -> str:
    return hashlib.sha256((s or "").encode("utf-8", errors="ignore")).hexdigest()

def read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")

def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")

def fmt_hms(seconds: float) -> str:
    seconds = max(0.0, float(seconds))
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    if h > 0:
        return f"{h}h {m:02d}m {s:02d}s"
    if m > 0:
        return f"{m}m {s:02d}s"
    return f"{s}s"

def _to01(x) -> str:
    x = "" if x is None else str(x).strip()
    return "1" if x in ("1", "1.0", "true", "True") else "0"

def resultado_texto(c1: str, c2: str) -> str:
    return "IDÊNTICO" if _to01(c1) == _to01(c2) else "DIFERENTE"

def is_all_zero_rows(rows: Dict[str, Dict[str, str]]) -> bool:
    for ph in PHASES:
        c1 = _to01(rows.get(ph, {}).get("check1", "0"))
        c2 = _to01(rows.get(ph, {}).get("check2", "0"))
        if c1 == "1" or c2 == "1":
            return False
    return True


# ============================================================
# Texto / vendor-only
# ============================================================

def limit_text(txt: str) -> str:
    txt = "" if txt is None else str(txt)
    lines = txt.splitlines()
    if SPIN_MAX_LINES_TOTAL > 0:
        lines = lines[:SPIN_MAX_LINES_TOTAL]
    out = "\n".join(lines).strip()
    if SPIN_MAX_CHARS_TOTAL > 0 and len(out) > SPIN_MAX_CHARS_TOTAL:
        out = out[:SPIN_MAX_CHARS_TOTAL].rstrip()
    return out

def extract_vendor_only(txt: str) -> str:
    txt = limit_text(txt)
    lines = txt.splitlines()

    vendor_lines: List[str] = []
    saw_vendor = False

    for line in lines:
        s = (line or "").strip()
        if not s:
            continue

        m = TAG_RE.match(s)
        if not m:
            continue

        head = (m.group(0) or "").strip().lower()
        content = TAG_RE.sub("", s).strip()
        if not content:
            continue

        if ("vendedor" in head) or ("agente" in head) or ("atendente" in head) or head.startswith("[vendedor]"):
            saw_vendor = True
            vendor_lines.append(f"[VENDEDOR] {content}")

    if saw_vendor and vendor_lines:
        return "\n".join(vendor_lines).strip()

    return txt.strip()


# ============================================================
# Prompt (sempre lido do arquivo)
# ============================================================

def load_prompt_file(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Prompt não encontrado: {path}")
    txt = read_text_file(path).strip()
    if not txt:
        raise RuntimeError(f"Prompt vazio: {path}")
    return txt

def pack_command_core(core: str, filename: str) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    core = (core or "")
    core = core.replace("{NOME_DO_ARQUIVO_ANEXADO}", filename)
    core = core.replace("{DATA_ANALISE}", today)
    return core.strip()

def build_prompt(core_packed: str, text_for_llm: str) -> str:
    text_for_llm = limit_text(text_for_llm)
    return (
        f"{core_packed}\n\n"
        f"[ANEXO — TRANSCRIÇÃO]\n"
        f"{text_for_llm}\n\n"
        f"[LEMBRETE FINAL — OBRIGATÓRIO]\n"
        f"Responda SOMENTE com as 6 linhas TSV no formato pedido (1 header + 5 linhas). "
        f"Não escreva explicações, títulos, listas ou texto extra."
    )


# ============================================================
# Ollama HTTP (com heartbeat)
# ============================================================

def _heartbeat(stop_evt: threading.Event, logger: logging.Logger, quiet: bool, prefix: str) -> None:
    t0 = time.time()
    # Primeira mensagem rápida para reforçar que está rodando
    if not quiet:
        logger.info(prefix + " iniciando…")
    while not stop_evt.wait(timeout=max(1, int(HEARTBEAT_EVERY_S))):
        if quiet:
            continue
        elapsed = int(time.time() - t0)
        logger.info(prefix + f" rodando há {elapsed}s…")

def call_ollama(prompt: str, timeout_s: int, logger: logging.Logger, quiet: bool = False) -> str:
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "keep_alive": OLLAMA_KEEP_ALIVE,
        "options": {
            "temperature": OLLAMA_TEMPERATURE,
            "num_ctx": OLLAMA_NUM_CTX,
            "num_predict": OLLAMA_NUM_PREDICT,
            "top_p": OLLAMA_TOP_P,
            "repeat_penalty": OLLAMA_REPEAT_PENALTY,
            "stop": ["```", "</s>"],
        },
    }

    data = json.dumps(payload).encode("utf-8")
    req = request.Request(
        OLLAMA_URL,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    last_err: Optional[Exception] = None

    for attempt in range(OLLAMA_TIMEOUT_RETRIES + 1):
        stop_evt = threading.Event()
        hb = threading.Thread(
            target=_heartbeat,
            args=(stop_evt, logger, quiet, f"Ollama ({OLLAMA_MODEL})"),
            daemon=True,
        )
        hb.start()

        try:
            with request.urlopen(req, timeout=timeout_s) as resp:
                raw = resp.read().decode("utf-8", errors="ignore")
            obj = json.loads(raw)
            out = (obj.get("response") or "").strip()
            return out
        except Exception as e:
            last_err = e
            if attempt < OLLAMA_TIMEOUT_RETRIES:
                if not quiet:
                    logger.info(f"Ollama falhou (tentativa {attempt+1}/{OLLAMA_TIMEOUT_RETRIES+1}). Repetindo…")
                time.sleep(2 + attempt * 2)
                continue
        finally:
            stop_evt.set()
            # não precisa join; thread é daemon e para com o event

    msg = str(last_err) if last_err else "unknown error"
    logger.error(f"Erro no Ollama: {msg}")
    raise RuntimeError(msg)


# ============================================================
# TSV Parse (TOLERANTE) + Canonicalização
# ============================================================

@dataclass
class TSVResult:
    ok: bool
    error: str
    raw_tsv: str          # o que guardamos no cache (canônico se ok; bruto se fail)
    table_rows: Dict[str, Dict[str, str]]

_SPLIT_RE = re.compile(r"\t+|\s{2,}|\s*\|\s*|;|,")

def _normalize_line(s: str) -> str:
    # remove aspas e caracteres invisíveis comuns
    s = (s or "").replace("\u200b", "").replace("\ufeff", "").strip()
    s = s.strip("`").strip()
    return s

def _looks_like_header(line: str) -> bool:
    l = line.lower()
    return ("spin" in l and "check" in l) or l.replace("_", " ").strip() in (
        "spin selling check 01 check 02",
        "spin selling check_01 check_02",
    )

def _parse_numbers_from_parts(parts: List[str]) -> Tuple[Optional[str], Optional[str]]:
    nums = []
    for p in parts:
        x = (p or "").strip()
        if x in ("0", "1"):
            nums.append(x)
        elif x.lower() in ("true", "false"):
            nums.append("1" if x.lower() == "true" else "0")
        elif x in ("1.0", "0.0"):
            nums.append("1" if x.startswith("1") else "0")
        if len(nums) >= 2:
            break
    if len(nums) >= 2:
        return nums[0], nums[1]
    return None, None

def canonicalize_tsv_and_rows(raw: str) -> Tuple[bool, str, str, Dict[str, Dict[str, str]]]:
    """
    Retorna:
      ok, err, canonical_tsv, rows
    Regras tolerantes:
      - aceita header com variações (tabs, espaços, underscores, pipes)
      - aceita saída com 5 linhas (sem header) desde que tenha P0..P4
      - aceita 4 colunas (com RESULTADO) e ignora colunas extra
      - aceita separador por TAB, múltiplos espaços, ou pipes
      - ignora linhas fora do padrão (ex: "Aqui está:", blocos, etc.)
    """
    if not raw or not str(raw).strip():
        return False, "empty_response", "", {}

    # filtra linhas úteis e tenta identificar as fases
    lines0 = [_normalize_line(x) for x in str(raw).splitlines()]
    lines = [ln for ln in lines0 if ln]

    if not lines:
        return False, "empty_response", "", {}

    rows: Dict[str, Dict[str, str]] = {ph: {"check1": "0", "check2": "0"} for ph in PHASES}
    got = set()

    for ln in lines:
        # pular headers e lixo comum
        if _looks_like_header(ln):
            continue
        lnl = ln.lower()
        if lnl.startswith(("aqui está", "segue", "resultado", "resposta", "tsv", "```")):
            continue

        # tenta achar qual fase é
        phase = None
        for ph in PHASES:
            if ln.startswith(ph):
                phase = ph
                rest = ln[len(ph):].strip()
                parts = _SPLIT_RE.split(rest)
                c1, c2 = _parse_numbers_from_parts(parts)
                if c1 is not None and c2 is not None:
                    rows[ph]["check1"] = c1
                    rows[ph]["check2"] = c2
                    got.add(ph)
                break

        if phase is not None:
            continue

        # tolerância: às vezes vem "P0_abertura 1 1 1" sem tabs
        for ph in PHASES:
            if ph in ln:
                # pega o pedaço após a ocorrência da fase
                idx = ln.find(ph)
                rest = ln[idx + len(ph):].strip()
                parts = _SPLIT_RE.split(rest)
                c1, c2 = _parse_numbers_from_parts(parts)
                if c1 is not None and c2 is not None:
                    rows[ph]["check1"] = c1
                    rows[ph]["check2"] = c2
                    got.add(ph)
                break

    if len(got) != len(PHASES):
        # Erro detalhado com fases faltantes (ajuda debug)
        missing = [ph for ph in PHASES if ph not in got]
        return False, f"missing_phases:{','.join(missing)}", "", {}

    # monta TSV canônico
    out_lines = [TSV_HEADER]
    for ph in PHASES:
        out_lines.append(f"{ph}\t{_to01(rows[ph]['check1'])}\t{_to01(rows[ph]['check2'])}")
        # normaliza para 0/1
        rows[ph]["check1"] = _to01(rows[ph]["check1"])
        rows[ph]["check2"] = _to01(rows[ph]["check2"])

    canonical = "\n".join(out_lines).strip()
    return True, "", canonical, rows


# ============================================================
# Excel
# ============================================================

def write_excel(out_xlsx: Path, transcript_title: str, table_rows: Dict[str, Dict[str, str]]) -> None:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18

    ws["A1"] = str(transcript_title)

    ws["A3"] = "SPIN SELLING"
    ws["B3"] = "CHECK_01"
    ws["C3"] = "CHECK_02"
    ws["D3"] = "RESULTADO TEXTO"

    r = 4
    for ph in PHASES:
        c1 = table_rows.get(ph, {}).get("check1", "0")
        c2 = table_rows.get(ph, {}).get("check2", "0")

        ws[f"A{r}"] = ph
        ws[f"B{r}"] = int(_to01(c1))
        ws[f"C{r}"] = int(_to01(c2))
        ws[f"D{r}"] = resultado_texto(c1, c2)
        r += 1

    wb.save(str(out_xlsx))


# ============================================================
# Cache SQLite
# ============================================================

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS cache (
  key TEXT PRIMARY KEY,
  text_sha256 TEXT NOT NULL,
  prompt_sha256 TEXT NOT NULL,
  model TEXT NOT NULL,
  created_at TEXT NOT NULL,
  status TEXT NOT NULL,         -- ok | fail
  tsv_raw TEXT NOT NULL,
  error TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_cache_text_sha256 ON cache(text_sha256);
"""

def cache_init(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    try:
        conn.executescript(SCHEMA_SQL)
        conn.commit()
    finally:
        conn.close()

def cache_get(db_path: Path, key: str) -> Optional[Dict[str, str]]:
    conn = sqlite3.connect(str(db_path))
    try:
        cur = conn.cursor()
        cur.execute("SELECT key, status, tsv_raw, error, created_at FROM cache WHERE key = ?", (key,))
        row = cur.fetchone()
        if not row:
            return None
        return {
            "key": row[0],
            "status": row[1],
            "tsv_raw": row[2],
            "error": row[3],
            "created_at": row[4],
        }
    finally:
        conn.close()

def cache_set(
    db_path: Path,
    key: str,
    text_sha256: str,
    prompt_sha256: str,
    model: str,
    status: str,
    tsv_raw: str,
    error: str,
) -> None:
    conn = sqlite3.connect(str(db_path))
    try:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO cache(key, text_sha256, prompt_sha256, model, created_at, status, tsv_raw, error)
            VALUES(?,?,?,?,?,?,?,?)
            ON CONFLICT(key) DO UPDATE SET
              text_sha256=excluded.text_sha256,
              prompt_sha256=excluded.prompt_sha256,
              model=excluded.model,
              created_at=excluded.created_at,
              status=excluded.status,
              tsv_raw=excluded.tsv_raw,
              error=excluded.error
            """,
            (key, text_sha256, prompt_sha256, model, now_iso(), status, tsv_raw, error),
        )
        conn.commit()
    finally:
        conn.close()


# ============================================================
# Arquivamento de TXT processado
# ============================================================

def safe_move_to_archive(in_path: Path, in_root: Path, logger: logging.Logger) -> None:
    """
    Move o TXT para ARCHIVE_DIR, preservando estrutura relativa ao in_root.
    Se já existir arquivo com mesmo nome, adiciona sufixo de timestamp.
    """
    try:
        rel = in_path.relative_to(in_root)
    except Exception:
        rel = Path(in_path.name)

    dest = ARCHIVE_DIR / rel
    dest.parent.mkdir(parents=True, exist_ok=True)

    if dest.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = dest.with_name(f"{dest.stem}_{ts}{dest.suffix}")

    try:
        shutil.move(str(in_path), str(dest))
    except Exception as e:
        logger.error(f"Falha ao mover para archive: {in_path} -> {dest} | {e}")


# ============================================================
# Processamento de arquivo
# ============================================================

@dataclass
class JobResult:
    ok: bool
    in_path: Path
    out_xlsx: Path
    used_cache: bool
    error: str

def build_cache_key(text_sha: str, prompt_sha: str, model: str, vendor_only: bool) -> str:
    base = f"spin02|v8_1_1|{model}|prompt={prompt_sha}|text={text_sha}|vendor_only={int(vendor_only)}"
    return sha256_text(base)

def run_once(core: str, filename: str, text_for_llm: str, logger: logging.Logger, quiet: bool) -> TSVResult:
    core_packed = pack_command_core(core, filename)
    prompt = build_prompt(core_packed, text_for_llm)

    t0 = time.time()
    raw = call_ollama(prompt, timeout_s=OLLAMA_TIMEOUT_S, logger=logger, quiet=quiet)
    dt = time.time() - t0
    if not quiet:
        logger.info(f"Ollama finalizou em {fmt_hms(dt)}.")

    ok, err, canonical, rows = canonicalize_tsv_and_rows(raw)
    if not ok:
        # em falha, guardamos o raw para debug
        return TSVResult(ok=False, error=err, raw_tsv=(raw or "").strip(), table_rows={})

    # em sucesso, guardamos o TSV canônico (estável para cache/replay)
    return TSVResult(ok=True, error="", raw_tsv=canonical, table_rows=rows)

def process_one(
    in_path: Path,
    in_root: Path,
    out_dir: Path,
    prompt_main: str,
    prompt_alt: str,
    prompt_sha256: str,
    db_path: Path,
    force: bool,
    quiet: bool,
    logger: logging.Logger,
) -> JobResult:
    stem = in_path.stem
    out_xlsx = out_dir / f"{stem}_SPIN.xlsx"

    # Lê TXT
    try:
        raw_txt = read_text_file(in_path)
    except Exception as e:
        logger.error(f"Falha ao ler TXT: {in_path} | {e}")
        rows_fail = {ph: {"check1": "0", "check2": "0"} for ph in PHASES}
        write_excel(out_xlsx, f"{stem} — FALHA", rows_fail)
        return JobResult(ok=False, in_path=in_path, out_xlsx=out_xlsx, used_cache=False, error=str(e))

    # Prepara texto
    text_for_llm = extract_vendor_only(raw_txt) if SPIN_VENDOR_ONLY else limit_text(raw_txt)
    text_sha = sha256_text(text_for_llm)
    cache_key = build_cache_key(text_sha, prompt_sha256, OLLAMA_MODEL, SPIN_VENDOR_ONLY)

    # Cache hit
    if not force:
        cached = cache_get(db_path, cache_key)
        if cached and cached.get("status") == "ok":
            tsv_raw = cached.get("tsv_raw", "")
            ok, err, canonical, rows = canonicalize_tsv_and_rows(tsv_raw)
            if ok:
                write_excel(out_xlsx, stem, rows)
                if not quiet:
                    logger.info(f"OK  | cache | {in_path.name} -> {out_xlsx.name}")
                safe_move_to_archive(in_path, in_root, logger)
                return JobResult(ok=True, in_path=in_path, out_xlsx=out_xlsx, used_cache=True, error="")

    # Execução principal
    err_msg = ""
    tsv_raw_best = ""
    rows_best: Dict[str, Dict[str, str]] = {ph: {"check1": "0", "check2": "0"} for ph in PHASES}

    try:
        res_main = run_once(prompt_main, in_path.name, text_for_llm, logger=logger, quiet=quiet)

        if res_main.ok:
            tsv_raw_best = res_main.raw_tsv
            rows_best = res_main.table_rows

            # Se ALL-ZERO, faz verificação secundária (silenciosa)
            if prompt_alt.strip() and is_all_zero_rows(rows_best):
                res_alt = run_once(prompt_alt, in_path.name, text_for_llm, logger=logger, quiet=True)
                if res_alt.ok and (not is_all_zero_rows(res_alt.table_rows)):
                    tsv_raw_best = res_alt.raw_tsv
                    rows_best = res_alt.table_rows

            cache_set(db_path, cache_key, text_sha, prompt_sha256, OLLAMA_MODEL, "ok", tsv_raw_best, "")
            write_excel(out_xlsx, stem, rows_best)

            if not quiet:
                logger.info(f"OK  | run   | {in_path.name} -> {out_xlsx.name}")

            safe_move_to_archive(in_path, in_root, logger)
            return JobResult(ok=True, in_path=in_path, out_xlsx=out_xlsx, used_cache=False, error="")

        # Se inválido, tenta alternativa (silenciosa)
        err_msg = res_main.error or "invalid_tsv"
        tsv_raw_best = res_main.raw_tsv or ""

        if prompt_alt.strip():
            res_alt = run_once(prompt_alt, in_path.name, text_for_llm, logger=logger, quiet=True)
            if res_alt.ok:
                tsv_raw_best = res_alt.raw_tsv
                rows_best = res_alt.table_rows

                cache_set(db_path, cache_key, text_sha, prompt_sha256, OLLAMA_MODEL, "ok", tsv_raw_best, "")
                write_excel(out_xlsx, stem, rows_best)

                if not quiet:
                    logger.info(f"OK  | run   | {in_path.name} -> {out_xlsx.name}")

                safe_move_to_archive(in_path, in_root, logger)
                return JobResult(ok=True, in_path=in_path, out_xlsx=out_xlsx, used_cache=False, error="")

            err_msg = res_alt.error or err_msg

    except Exception as e:
        err_msg = str(e) or err_msg

    # Falha final
    logger.error(f"FAIL| {in_path} | {err_msg}")
    cache_set(db_path, cache_key, text_sha, prompt_sha256, OLLAMA_MODEL, "fail", (tsv_raw_best or ""), err_msg)
    write_excel(out_xlsx, f"{stem} — FALHA", rows_best)
    safe_move_to_archive(in_path, in_root, logger)
    return JobResult(ok=False, in_path=in_path, out_xlsx=out_xlsx, used_cache=False, error=err_msg)


# ============================================================
# Discovery de arquivos
# ============================================================

def discover_txts(in_dir: Path, pattern: str, recursive: bool) -> List[Path]:
    if recursive:
        return sorted([p for p in in_dir.rglob(pattern) if p.is_file()])
    return sorted([p for p in in_dir.glob(pattern) if p.is_file()])


# ============================================================
# CLI
# ============================================================

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="SPIN Analyzer 02 (LOCAL) — TXT -> Ollama -> TSV -> Excel",
    )
    p.add_argument("--in_dir", default=str(DEFAULT_IN_DIR), help="Pasta de entrada (default: arquivos_transcritos/txt)")
    p.add_argument("--out_dir", default=str(DEFAULT_OUT_DIR), help="Pasta de saída (default: saida_excel)")
    p.add_argument("--pattern", default="*.txt", help="Padrão de arquivos (default: *.txt)")
    p.add_argument("--recursive", default="true", help="true/false (default: true)")
    p.add_argument("--workers", type=int, default=1, help="Número de workers (default: 1)")
    p.add_argument("--force", action="store_true", help="Ignora cache e reprocessa")
    p.add_argument("--quiet", action="store_true", help="Reduz logs no console")
    return p.parse_args()

def parse_bool(s: str, default: bool = True) -> bool:
    if s is None:
        return default
    v = str(s).strip().lower()
    if v in ("1", "true", "yes", "y", "sim"):
        return True
    if v in ("0", "false", "no", "n", "nao", "não"):
        return False
    return default


# ============================================================
# Main
# ============================================================

def main() -> int:
    args = parse_args()
    quiet = bool(args.quiet)
    logger = setup_logging(quiet=quiet)

    in_dir = Path(args.in_dir).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    # cria archive dir
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

    recursive = parse_bool(args.recursive, default=True)
    workers = max(1, int(args.workers or 1))
    force = bool(args.force)

    # Prompts sempre do arquivo
    try:
        prompt_main = load_prompt_file(PROMPT_MAIN_PATH)
        prompt_alt = load_prompt_file(PROMPT_ALT_PATH) if PROMPT_ALT_PATH.exists() else ""
    except Exception as e:
        logger.error(f"Falha ao carregar prompt(s) em assets/: {e}")
        return 1

    prompt_sha = sha256_text(prompt_main)

    # Cache DB init
    cache_init(CACHE_DB_PATH)

    if not in_dir.exists() or not in_dir.is_dir():
        logger.error(f"Pasta de entrada inválida: {in_dir}")
        return 1

    files = discover_txts(in_dir, args.pattern, recursive=recursive)
    if not files:
        logger.info("Nenhum TXT encontrado.")
        return 0

    if not quiet:
        logger.info("SPIN Analyzer 02 (LOCAL)")
        logger.info(f"Início: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"IN : {in_dir}")
        logger.info(f"OUT: {out_dir}")
        logger.info(f"Modelo: {OLLAMA_MODEL}")
        logger.info(f"Vendor-only: {'SIM' if SPIN_VENDOR_ONLY else 'NÃO'}")
        logger.info(f"Arquivos: {len(files)}")

    t0 = time.time()
    failed = 0

    if workers == 1:
        for i, fp in enumerate(files, start=1):
            t_file = time.time()

            if not quiet:
                logger.info(f"Processando: {fp.name}")

            res = process_one(
                in_path=fp,
                in_root=in_dir,
                out_dir=out_dir,
                prompt_main=prompt_main,
                prompt_alt=prompt_alt,
                prompt_sha256=prompt_sha,
                db_path=CACHE_DB_PATH,
                force=force,
                quiet=quiet,
                logger=logger,
            )

            if not res.ok:
                failed += 1

            if not quiet:
                dt = time.time() - t_file
                done = i
                total = len(files)
                if done >= 1:
                    avg = (time.time() - t0) / max(1, done)
                    eta = (total - done) * avg
                    logger.info(f"Progresso: {done}/{total} | Último: {fmt_hms(dt)} | ETA: {fmt_hms(eta)}")

    else:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        with ThreadPoolExecutor(max_workers=workers) as ex:
            futs = []
            for fp in files:
                futs.append(ex.submit(
                    process_one,
                    fp, in_dir, out_dir, prompt_main, prompt_alt, prompt_sha,
                    CACHE_DB_PATH, force, quiet, logger
                ))

            done = 0
            total = len(futs)
            for fut in as_completed(futs):
                done += 1
                res = fut.result()
                if not res.ok:
                    failed += 1
                if not quiet:
                    avg = (time.time() - t0) / max(1, done)
                    eta = (total - done) * avg
                    logger.info(f"Progresso: {done}/{total} | ETA: {fmt_hms(eta)}")

    total_s = time.time() - t0
    if not quiet:
        logger.info(f"Fim: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"Duração total: {fmt_hms(total_s)}")

    return 1 if failed > 0 else 0


if __name__ == "__main__":
    raise SystemExit(main())
